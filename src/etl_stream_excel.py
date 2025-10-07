import os
from collections import defaultdict
from typing import Dict, Iterable, List, Tuple, Optional

import pandas as pd
from openpyxl import load_workbook
from tqdm.auto import tqdm
import psutil

from .schema import fuzzy_detect, CANON
from .io_utils import ensure_output_dirs, write_parquet
from .logger import get_logger

def _to_year(v):
    import datetime, re
    if v is None or v == "":
        return None
    if isinstance(v, (datetime.date, datetime.datetime)):
        return int(v.year)
    s = str(v).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y/%m/%d", "%m/%d/%y"):
        try:
            return datetime.datetime.strptime(s, fmt).year
        except Exception:
            pass
    m = re.search(r"(19|20)\d{2}", s)
    return int(m.group(0)) if m else None


def _mem_mb() -> float:
	process = psutil.Process()
	return process.memory_info().rss / (1024 * 1024)


def detect_columns(header_row: List[str]) -> Dict[str, int]:
    mapping = fuzzy_detect(header_row)
    logger = get_logger()
    selected = {}
    for k in ["country", "sector", "year", "patent_id"]:
        if k not in mapping:
            raise ValueError(
                f"Required column '{k}' not found. Detected={mapping}. Hint: rename headers or provide mapping."
            )
        selected[k] = mapping[k]
    # quality optional
    if "quality" in mapping:
        selected["quality"] = mapping["quality"]

    for proxy_key in ["nb_citing_docdb_fam", "docdb_family_size", "granted"]:
        if proxy_key in mapping:
            selected[proxy_key] = mapping[proxy_key]

    logger.info(f"Header detection: {selected}")
    return selected



def _normalize_country(val: str) -> Optional[str]:
	if val is None:
		return None
	s = str(val).strip()
	if not s:
		return None
	return s.upper() if len(s) in (2, 3) else s.title()


def _normalize_sector(val: str) -> Optional[str]:
	if val is None:
		return None
	s = str(val).strip()
	return s if s else None


def _to_int(val) -> Optional[int]:
	try:
		iv = int(val)
		return iv
	except Exception:
		return None


def _flush_part(agg: Dict[Tuple[str, str, int], Dict[str, float]], out_dir: str, part_idx: int) -> int:
	if not agg:
		return 0
	rows = []
	for (c, s, y), v in agg.items():
		q = v.get("Q_sum") / v.get("Q_n") if v.get("Q_n", 0) > 0 else None
		rows.append({
			CANON["country"]: c,
			CANON["sector"]: s,
			CANON["year"]: y,
			"P": v.get("P", 0),
			"Q": q,
		})
	df = pd.DataFrame(rows)
	path = os.path.join(out_dir, "tmp_parts", f"part_{part_idx:05d}.parquet")
	write_parquet(df, path)
	return len(df)


def _concat_parts(out_dir: str) -> str:
	parts_dir = os.path.join(out_dir, "tmp_parts")
	paths = sorted([os.path.join(parts_dir, p) for p in os.listdir(parts_dir) if p.endswith('.parquet')])
	if not paths:
		raise FileNotFoundError("No parquet parts to concatenate.")
	dfs = [pd.read_parquet(p) for p in paths]
	full = pd.concat(dfs, ignore_index=True)
	full = (
		full.groupby([CANON["country"], CANON["sector"], CANON["year"]], as_index=False)
		.agg({"P": "sum", "Q": "mean"})
	)
	final_path = os.path.join(out_dir, "cst_agg.parquet")
	write_parquet(full, final_path)
	return final_path


def stream_aggregate_xlsx(path: str) -> str:
    logger = get_logger()
    out_dir = os.getenv("OUT_DIR", "out")
    ensure_output_dirs(out_dir)
    if not os.path.exists(path):
        raise FileNotFoundError(f"Input Excel not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ValueError("Empty Excel sheet")

    col_idx = detect_columns(list(header))

    agg: Dict[Tuple[str, str, int], Dict[str, float]] = defaultdict(lambda: {"P": 0, "Q_sum": 0.0, "Q_n": 0})
    part_idx = 0
    rows_since_flush = 0
    flush_every = int(os.getenv("PART_SIZE_ROWS", "200000"))
    max_groups = int(os.getenv("FLUSH_MAX_GROUPS", "500000"))

    # float tool
    def _safe_float(x):
        try:
            return float(str(x).replace(",", "")) if x is not None else 0.0
        except Exception:
            return 0.0

    for row in tqdm(rows_iter, desc="Streaming rows"):
        country = _normalize_country(row[col_idx["country"]])
        sector  = _normalize_sector(row[col_idx["sector"]])
        year    = _to_year(row[col_idx["year"]])

        if not (country and sector and year):
            continue

        key = (country, sector, year)
        entry = agg[key]
        entry["P"] += 1

        # quality if not proxy 
        if "quality" in col_idx:
            q_val = row[col_idx["quality"]]
            try:
                q = float(q_val) if q_val is not None else None
            except Exception:
                q = None
            if q is not None:
                entry["Q_sum"] += q
                entry["Q_n"] += 1
        else:
            cites = _safe_float(row[col_idx["nb_citing_docdb_fam"]]) if "nb_citing_docdb_fam" in col_idx else 0.0
            fam   = _safe_float(row[col_idx["docdb_family_size"]])  if "docdb_family_size"  in col_idx else 0.0
            granted_flag = row[col_idx["granted"]] if "granted" in col_idx else ""
            granted_bin = 1.0 if str(granted_flag).strip().upper() in ("Y", "YES", "TRUE", "1") else 0.0
            q_proxy = 0.6 * cites + 0.3 * fam + 0.1 * granted_bin
            entry["Q_sum"] += q_proxy
            entry["Q_n"]   += 1

        rows_since_flush += 1
        if rows_since_flush >= flush_every or len(agg) >= max_groups:
            count = _flush_part(agg, out_dir, part_idx)
            logger.info(f"Flushed part {part_idx} with {count} rows; mem={_mem_mb():.1f} MB; groups={len(agg)}")
            agg.clear()
            part_idx += 1
            rows_since_flush = 0

    # final flush
    count = _flush_part(agg, out_dir, part_idx)
    logger.info(f"Final flush part {part_idx} with {count} rows; mem={_mem_mb():.1f} MB")
    agg.clear()

    final_path = _concat_parts(out_dir)
    logger.info(f"Wrote aggregated parquet: {final_path}")
    return final_path

